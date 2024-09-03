import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.worksheet.table import Table, TableStyleInfo
import xlwings as xw
import time
import os

t1 = time.time()

base_file_location = r"C:\Users\aasabu\Desktop\Sales and Marketing\Python codes\Requirement Subclass"
reqDump_file_location = base_file_location + r"\Req_US_Dump.xlsx"
output_location = base_file_location + r"\ReqSubClassOutput.xlsx"

reqDump = pd.read_excel(reqDump_file_location)

gapSubClasses = ['Gap - Process Deviation', 'Gap - WRICEF-(E)-Enhancement', 'Gap - Extended configuration not aligned to Best Practices']

def check_sub_class(row,array):

    if str(row['SubClassification']) == str(row['SubClassification2']):
        array.append(str(row['Requirement ID']))
        return 'True'
    return ''

def gap_analysis(row,array):

    if str(row['SubClassification']) == 'Gap - Process Deviation' or str(row['SubClassification']) == 'Gap - WRICEF-(E)-Enhancement' or str(row['SubClassification']) == 'Gap - Extended configuration not aligned to Best Practices':
        
        if str(row['SubClassification']) != str(row['SubClassification2']):
            for x in array:
                if str(row['Requirement ID']) == x:
                    return ''
            return 'True'
            #return 'No US Subclass matches req subclass'
        
    return ''

def gap_check(row):
    if str(row['GAP Req Check']) == 'True' and (str(row['SubClassification2']) == 'Gap - Process Deviation' or str(row['SubClassification2']) == 'Gap - WRICEF-(E)-Enhancement' or str(row['SubClassification2']) == 'Gap - Extended configuration not aligned to Best Practices'):
        return 'GAP Req - No 1:1 match with GAP US'
    return ''

req_Array = []

reqDump['Check'] = reqDump.apply(check_sub_class,args=(req_Array,),axis=1)

reqDump1 = reqDump[reqDump['User Story Title'].notnull()]
reqDump2 = reqDump1[reqDump1['User Story State'] != 'Removed'].copy()

gapReqDF = reqDump2[reqDump2['SubClassification'] .isin(gapSubClasses)].copy()


gapReqDF['GAP Req Check'] = gapReqDF.apply(gap_analysis,args=(req_Array,),axis=1)
gapReqDF['GAP?'] = gapReqDF.apply(gap_check, axis=1)
#reqDump['Fit-GAP Issues'] = reqDump.apply(fit_to_gap_check,axis=1)
'''
reqDump['FIT - 1:1 Match'] = ''
reqDump['FIT - atleast 1 Match'] = ''
reqDump['FIT - no match'] = ''

for req_id in reqDump['Requirement ID'].unique():

    req_group= reqDump[reqDump['Requirement ID'] == req_id]

    if req_group['GAP Req Check'].any():
        continue
    if req_group['Fit-GAP Issues'].any():
        continue
    if req_group['Check'].all():
        reqDump.loc[reqDump['Requirement ID'] == req_id, 'FIT - 1:1 Match'] = 'True'
    elif req_group['Check'].any():
        reqDump.loc[reqDump['Requirement ID'] == req_id, 'FIT - atleast 1 Match'] = 'True'
    else:
        reqDump.loc[reqDump['Requirement ID'] == req_id, 'FIT - no match'] = 'True'

print(req_Array)
'''

with pd.ExcelWriter(output_location, engine='openpyxl') as writer:
    gapReqDF.to_excel(writer, sheet_name='BaseFile', index=False)

print("Converting to table...")
#Applying table formatting
wb = load_workbook(output_location)
sheet1 = wb['BaseFile']

sheetLists = [sheet1]
x = 1

for sheetList in sheetLists: 
    
    min_row = sheetList.min_row
    min_col = sheetList.min_column
    max_row = sheetList.max_row
    max_col = sheetList.max_column

    startCell = sheetList.cell(row=min_row,column = min_col).coordinate
    endCell = sheetList.cell(row=max_row,column = max_col).coordinate

    table_ref = f"{startCell}:{endCell}"
    
    table_name = 'Table'+str(x)
    x = x+1
    table = Table(displayName=table_name,ref = table_ref)

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )

    table.tableStyleInfo = style
    sheetList.add_table(table)

    wb.save(output_location)   

vba_code = """

Sub Macro2()
'
' Macro2 Macro
'

'
Dim sheetname As String

    Sheets.Add
    sheetname = ActiveSheet.Name
    Sheets(sheetname).Select
    Sheets(sheetname).Name = "Overview"
    Sheets("Overview").Select
    
    ActiveCell.Offset(12, 1).Range("A1").Select
    Workbooks("ReqSubClassOutput.xlsx").Connections.Add2 _
        "WorksheetConnection_ReqSubClassOutput.xlsx!Table1", "", _
        "WORKSHEET;C:\\Users\\aasabu\\Desktop\\Sales and Marketing\\Python codes\Requirement Subclass\\ReqSubClassOutput.xlsx" _
        , "ReqSubClassOutput.xlsx!Table1", 7, True, False
    ActiveWorkbook.PivotCaches.Create(SourceType:=xlExternal, SourceData:= _
        ActiveWorkbook.Connections( _
        "WorksheetConnection_ReqSubClassOutput.xlsx!Table1"), Version:=8). _
        CreatePivotTable TableDestination:="Overview!R13C2", TableName:= _
        "PivotTable1", DefaultVersion:=8
    Cells(13, 2).Select
    With ActiveSheet.PivotTables("PivotTable1")
        .ColumnGrand = True
        .HasAutoFormat = True
        .DisplayErrorString = False
        .DisplayNullString = True
        .EnableDrilldown = True
        .ErrorString = ""
        .MergeLabels = False
        .NullString = ""
        .PageFieldOrder = 2
        .PageFieldWrapCount = 0
        .PreserveFormatting = True
        .RowGrand = True
        .PrintTitles = False
        .RepeatItemsOnEachPrintedPage = True
        .TotalsAnnotation = True
        .CompactRowIndent = 1
        .VisualTotals = False
        .InGridDropZones = False
        .DisplayFieldCaptions = True
        .DisplayMemberPropertyTooltips = True
        .DisplayContextTooltips = True
        .ShowDrillIndicators = True
        .PrintDrillIndicators = False
        .DisplayEmptyRow = False
        .DisplayEmptyColumn = False
        .AllowMultipleFilters = False
        .SortUsingCustomLists = True
        .DisplayImmediateItems = True
        .ViewCalculatedMembers = True
        .FieldListSortAscending = False
        .ShowValuesRow = False
        .CalculatedMembersInFilters = True
        .RowAxisLayout xlCompactRow
    End With
    ActiveSheet.PivotTables("PivotTable1").PivotCache.RefreshOnFileOpen = False
    ActiveSheet.PivotTables("PivotTable1").RepeatAllLabels xlRepeatLabels
    ActiveWindow.SmallScroll Down:=3
    ActiveSheet.PivotTables("PivotTable1").CubeFields.GetMeasure _
        "[Table1].[Requirement Title]", xlCount, "Count of Requirement Title"
    ActiveSheet.PivotTables("PivotTable1").AddDataField ActiveSheet.PivotTables( _
        "PivotTable1").CubeFields("[Measures].[Count of Requirement Title]"), _
        "Count of Requirement Title"
    With ActiveSheet.PivotTables("PivotTable1").CubeFields( _
        "[Table1].[SubClassification]")
        .Orientation = xlRowField
        .Position = 1
    End With
    With ActiveSheet.PivotTables("PivotTable1").CubeFields( _
        "[Table1].[GAP Req Check]")
        .Orientation = xlPageField
        .Position = 1
    End With
    ActiveSheet.PivotTables("PivotTable1").CubeFields(15).EnableMultiplePageItems _
        = True
    ActiveSheet.PivotTables("PivotTable1").PivotFields( _
        "[Table1].[GAP Req Check].[GAP Req Check]").VisibleItemsList = Array( _
        "[Table1].[GAP Req Check].&")
    ActiveCell.Offset(0, 1).Range("A1").Select
    ActiveSheet.PivotTables("PivotTable1").PivotFields( _
        "[Measures].[Count of Requirement Title]").Function = xlDistinctCount
    
    ActiveCell.Offset(0, -1).Range("A1").Select
    ActiveSheet.PivotTables("PivotTable1").CompactLayoutRowHeader = _
        "GAP Req - GAP US 1:1 Match"
    
    ActiveWindow.Zoom = 85
    ActiveWindow.Zoom = 70
    ActiveCell.Offset(0, 3).Range("A1").Select

    ActiveSheet.PivotTables("PivotTable1").TableStyle2 = "PivotStyleLight17"
    ActiveWindow.DisplayGridlines = False

    ActiveWorkbook.PivotCaches.Create(SourceType:=xlExternal, SourceData:= _
        ActiveWorkbook.Connections( _
        "WorksheetConnection_ReqSubClassOutput.xlsx!Table1"), Version:=8). _
        CreatePivotTable TableDestination:="Overview!R13C5", TableName:= _
        "PivotTable2", DefaultVersion:=8
    Cells(13, 5).Select
    With ActiveSheet.PivotTables("PivotTable2")
        .ColumnGrand = True
        .HasAutoFormat = True
        .DisplayErrorString = False
        .DisplayNullString = True
        .EnableDrilldown = True
        .ErrorString = ""
        .MergeLabels = False
        .NullString = ""
        .PageFieldOrder = 2
        .PageFieldWrapCount = 0
        .PreserveFormatting = True
        .RowGrand = True
        .PrintTitles = False
        .RepeatItemsOnEachPrintedPage = True
        .TotalsAnnotation = True
        .CompactRowIndent = 1
        .VisualTotals = False
        .InGridDropZones = False
        .DisplayFieldCaptions = True
        .DisplayMemberPropertyTooltips = True
        .DisplayContextTooltips = True
        .ShowDrillIndicators = True
        .PrintDrillIndicators = False
        .DisplayEmptyRow = False
        .DisplayEmptyColumn = False
        .AllowMultipleFilters = False
        .SortUsingCustomLists = True
        .DisplayImmediateItems = True
        .ViewCalculatedMembers = True
        .FieldListSortAscending = False
        .ShowValuesRow = False
        .CalculatedMembersInFilters = True
        .RowAxisLayout xlCompactRow
    End With
    ActiveSheet.PivotTables("PivotTable2").PivotCache.RefreshOnFileOpen = False
    ActiveSheet.PivotTables("PivotTable2").RepeatAllLabels xlRepeatLabels
    ActiveSheet.PivotTables("PivotTable2").AddDataField ActiveSheet.PivotTables( _
        "PivotTable2").CubeFields("[Measures].[Count of Requirement Title]"), _
        "Count of Requirement Title"
    With ActiveSheet.PivotTables("PivotTable2").CubeFields( _
        "[Table1].[SubClassification]")
        .Orientation = xlRowField
        .Position = 1
    End With
    With ActiveSheet.PivotTables("PivotTable2").CubeFields( _
        "[Table1].[GAP Req Check]")
        .Orientation = xlPageField
        .Position = 1
    End With
    ActiveCell.Offset(0, 1).Range("A1").Select
    ActiveSheet.PivotTables("PivotTable2").PivotFields( _
        "[Measures].[Count of Requirement Title]").Function = xlDistinctCount
    ActiveWindow.ScrollColumn = 2
    ActiveSheet.PivotTables("PivotTable2").CubeFields(15).EnableMultiplePageItems _
        = True
    ActiveSheet.PivotTables("PivotTable2").PivotFields( _
        "[Table1].[GAP Req Check].[GAP Req Check]").VisibleItemsList = Array( _
        "[Table1].[GAP Req Check].&[True]")
    
    
    
    ActiveCell.Offset(-11, 3).Range("A1").Select
    ActiveWorkbook.PivotCaches.Create(SourceType:=xlExternal, SourceData:= _
        ActiveWorkbook.Connections("WorksheetConnection_ReqSubClassOutput.xlsx!Table1") _
        , Version:=8).CreatePivotTable TableDestination:="Overview!R13C8", _
        TableName:="PivotTable16", DefaultVersion:=8
    Cells(13, 8).Select
    With ActiveSheet.PivotTables("PivotTable16")
        .ColumnGrand = True
        .HasAutoFormat = True
        .DisplayErrorString = False
        .DisplayNullString = True
        .EnableDrilldown = True
        .ErrorString = ""
        .MergeLabels = False
        .NullString = ""
        .PageFieldOrder = 2
        .PageFieldWrapCount = 0
        .PreserveFormatting = True
        .RowGrand = True
        .PrintTitles = False
        .RepeatItemsOnEachPrintedPage = True
        .TotalsAnnotation = True
        .CompactRowIndent = 1
        .VisualTotals = False
        .InGridDropZones = False
        .DisplayFieldCaptions = True
        .DisplayMemberPropertyTooltips = True
        .DisplayContextTooltips = True
        .ShowDrillIndicators = True
        .PrintDrillIndicators = False
        .DisplayEmptyRow = False
        .DisplayEmptyColumn = False
        .AllowMultipleFilters = False
        .SortUsingCustomLists = True
        .DisplayImmediateItems = True
        .ViewCalculatedMembers = True
        .FieldListSortAscending = False
        .ShowValuesRow = False
        .CalculatedMembersInFilters = True
        .RowAxisLayout xlCompactRow
    End With
    ActiveSheet.PivotTables("PivotTable16").PivotCache.RefreshOnFileOpen = False
    ActiveSheet.PivotTables("PivotTable16").RepeatAllLabels xlRepeatLabels
    ActiveSheet.PivotTables("PivotTable16").AddDataField ActiveSheet.PivotTables( _
        "PivotTable16").CubeFields("[Measures].[Count of Requirement Title]"), _
        "Count of Requirement Title"
    With ActiveSheet.PivotTables("PivotTable16").CubeFields( _
        "[Table1].[SubClassification]")
        .Orientation = xlRowField
        .Position = 1
    End With
    With ActiveSheet.PivotTables("PivotTable16").CubeFields( _
        "[Table1].[GAP Req Check]")
        .Orientation = xlPageField
        .Position = 1
    End With
    ActiveSheet.PivotTables("PivotTable16").CubeFields(15).EnableMultiplePageItems _
        = True
    ActiveSheet.PivotTables("PivotTable16").PivotFields( _
        "[Table1].[GAP Req Check].[GAP Req Check]").VisibleItemsList = Array( _
        "[Table1].[GAP Req Check].&[True]")
    With ActiveSheet.PivotTables("PivotTable16").CubeFields( _
        "[Table1].[SubClassification2]")
        .Orientation = xlRowField
        .Position = 2
    End With
    
    ActiveWorkbook.PivotCaches.Create(SourceType:=xlExternal, SourceData:= _
        ActiveWorkbook.Connections("WorksheetConnection_ReqSubClassOutput.xlsx!Table1") _
        , Version:=8).CreatePivotTable TableDestination:="Overview!R24C2", _
        TableName:="PivotTable13", DefaultVersion:=8
    Cells(24, 2).Select
    With ActiveSheet.PivotTables("PivotTable13")
        .ColumnGrand = True
        .HasAutoFormat = True
        .DisplayErrorString = False
        .DisplayNullString = True
        .EnableDrilldown = True
        .ErrorString = ""
        .MergeLabels = False
        .NullString = ""
        .PageFieldOrder = 2
        .PageFieldWrapCount = 0
        .PreserveFormatting = True
        .RowGrand = True
        .PrintTitles = False
        .RepeatItemsOnEachPrintedPage = True
        .TotalsAnnotation = True
        .CompactRowIndent = 1
        .VisualTotals = False
        .InGridDropZones = False
        .DisplayFieldCaptions = True
        .DisplayMemberPropertyTooltips = True
        .DisplayContextTooltips = True
        .ShowDrillIndicators = True
        .PrintDrillIndicators = False
        .DisplayEmptyRow = False
        .DisplayEmptyColumn = False
        .AllowMultipleFilters = False
        .SortUsingCustomLists = True
        .DisplayImmediateItems = True
        .ViewCalculatedMembers = True
        .FieldListSortAscending = False
        .ShowValuesRow = False
        .CalculatedMembersInFilters = True
        .RowAxisLayout xlCompactRow
    End With
    ActiveSheet.PivotTables("PivotTable13").PivotCache.RefreshOnFileOpen = False
    ActiveSheet.PivotTables("PivotTable13").RepeatAllLabels xlRepeatLabels
    ActiveSheet.PivotTables("PivotTable13").AddDataField ActiveSheet.PivotTables( _
        "PivotTable13").CubeFields("[Measures].[Count of Requirement Title]"), _
        "Count of Requirement Title"
    With ActiveSheet.PivotTables("PivotTable13").CubeFields( _
        "[Table1].[SubClassification]")
        .Orientation = xlRowField
        .Position = 1
    End With
    With ActiveSheet.PivotTables("PivotTable13").CubeFields("[Table1].[GAP?]")
        .Orientation = xlPageField
        .Position = 1
    End With
    ActiveSheet.PivotTables("PivotTable13").CubeFields(16).EnableMultiplePageItems _
        = True
    ActiveSheet.PivotTables("PivotTable13").PivotFields("[Table1].[GAP?].[GAP?]"). _
        VisibleItemsList = Array( _
        "[Table1].[GAP?].&[GAP Req - No 1:1 match with GAP US]")
    ActiveCell.Offset(7, 0).Range("A1").Select
    ActiveWorkbook.PivotCaches.Create(SourceType:=xlDatabase, SourceData:= _
        "Table1", Version:=8).CreatePivotTable TableDestination:="Overview!R31C2", _
        TableName:="PivotTable14", DefaultVersion:=8
    Sheets("Overview").Select
    Cells(31, 2).Select
    With ActiveSheet.PivotTables("PivotTable14")
        .ColumnGrand = True
        .HasAutoFormat = True
        .DisplayErrorString = False
        .DisplayNullString = True
        .EnableDrilldown = True
        .ErrorString = ""
        .MergeLabels = False
        .NullString = ""
        .PageFieldOrder = 2
        .PageFieldWrapCount = 0
        .PreserveFormatting = True
        .RowGrand = True
        .SaveData = True
        .PrintTitles = False
        .RepeatItemsOnEachPrintedPage = True
        .TotalsAnnotation = False
        .CompactRowIndent = 1
        .InGridDropZones = False
        .DisplayFieldCaptions = True
        .DisplayMemberPropertyTooltips = False
        .DisplayContextTooltips = True
        .ShowDrillIndicators = True
        .PrintDrillIndicators = False
        .AllowMultipleFilters = False
        .SortUsingCustomLists = True
        .FieldListSortAscending = False
        .ShowValuesRow = False
        .CalculatedMembersInFilters = False
        .RowAxisLayout xlCompactRow
    End With
    With ActiveSheet.PivotTables("PivotTable14").PivotCache
        .RefreshOnFileOpen = False
        .MissingItemsLimit = xlMissingItemsDefault
    End With
    ActiveSheet.PivotTables("PivotTable14").RepeatAllLabels xlRepeatLabels
    ActiveSheet.PivotTables("PivotTable14").AddDataField ActiveSheet.PivotTables( _
        "PivotTable14").PivotFields("User Story Title"), "Count of User Story Title", _
        xlCount
    With ActiveSheet.PivotTables("PivotTable14").PivotFields("SubClassification")
        .Orientation = xlRowField
        .Position = 1
    End With
    With ActiveSheet.PivotTables("PivotTable14").PivotFields("SubClassification2")
        .Orientation = xlRowField
        .Position = 2
    End With
    With ActiveSheet.PivotTables("PivotTable14").PivotFields("GAP?")
        .Orientation = xlPageField
        .Position = 1
    End With
    ActiveSheet.PivotTables("PivotTable14").PivotFields("GAP?").CurrentPage = _
        "(All)"
    With ActiveSheet.PivotTables("PivotTable14").PivotFields("GAP?")
        .PivotItems("(blank)").Visible = False
    End With
    ActiveSheet.PivotTables("PivotTable14").PivotFields("GAP?"). _
        EnableMultiplePageItems = True
    
    ActiveSheet.Range("B10").Select
    ActiveCell.FormulaR1C1 = "Case 1: GAP Requirements - GAP User Stories (Atleast one related US Subclass 1:1 matches with GAP Req)"
    ActiveCell.Select
    Selection.Font.Bold = True
    With Selection.Interior
        .Pattern = xlSolid
        .PatternColorIndex = xlAutomatic
        .Color = 65535
        .TintAndShade = 0
        .PatternTintAndShade = 0
    End With

    ActiveSheet.Range("B20").Select
    ActiveCell.FormulaR1C1 = "Case 3: GAP Req. without 1:1 match with GAP US (None of the related US subclass matches)"
    ActiveCell.Select
    Selection.Font.Bold = True
    With Selection.Interior
        .Pattern = xlSolid
        .PatternColorIndex = xlAutomatic
        .Color = 65535
        .TintAndShade = 0
        .PatternTintAndShade = 0
    End With
    
    ActiveSheet.Range("B28").Select
    ActiveCell.FormulaR1C1 = "Case 3.a: GAP US without 1:1 match with GAP Req (None of the related US subclass matches)"
    ActiveCell.Select
    Selection.Font.Bold = True
    With Selection.Interior
        .Pattern = xlSolid
        .PatternColorIndex = xlAutomatic
        .Color = 65535
        .TintAndShade = 0
        .PatternTintAndShade = 0
    End With
    
    ActiveSheet.Range("H10").Select
    ActiveCell.FormulaR1C1 = "Case 2.a: GAP US - FIT Req (None of the related US Subclass matches with GAP Req)"
    ActiveCell.Select
    Selection.Font.Bold = True
    With Selection.Interior
        .Pattern = xlSolid
        .PatternColorIndex = xlAutomatic
        .Color = 65535
        .TintAndShade = 0
        .PatternTintAndShade = 0
    End With

    ActiveSheet.Range("E10").Select
    ActiveCell.FormulaR1C1 = "Case 2: GAP Requirements without GAP US (None of related US is GAP)"
    ActiveCell.Select
    Selection.Font.Bold = True
    With Selection.Interior
        .Pattern = xlSolid
        .PatternColorIndex = xlAutomatic
        .Color = 65535
        .TintAndShade = 0
        .PatternTintAndShade = 0
    End With

    ActiveSheet.Range("B2:B3").Select
    ActiveCell.FormulaR1C1 = "GAP Requirements - Fit/Gap User Stories"
    
    With Selection.Interior
        .Pattern = xlSolid
        .PatternColorIndex = xlAutomatic
        .ThemeColor = xlThemeColorAccent3
        .TintAndShade = 0.399975585192419
        .PatternTintAndShade = 0
    End With
    Selection.Font.Bold = True
    
    ActiveSheet.Range("B2:B3").Select
    With Selection
        .HorizontalAlignment = xlCenter
        .VerticalAlignment = xlCenter
        .WrapText = False
        .Orientation = 0
        .AddIndent = False
        .IndentLevel = 0
        .ShrinkToFit = False
        .ReadingOrder = xlContext
        .MergeCells = True
    End With
    
    ActiveSheet.Range("C2:E3").Select
    ActiveCell.FormulaR1C1 = "GAP Requirements (i.e. Enhancements, Extended Config, Process Deviations) should have 1:1 match with atleast one of related US GAP Subclassifications"
    With Selection
        .HorizontalAlignment = xlCenter
        .VerticalAlignment = xlCenter
        .WrapText = True
        .Orientation = 0
        .AddIndent = False
        .IndentLevel = 0
        .ShrinkToFit = False
        .ReadingOrder = xlContext
        .MergeCells = True
    End With

    
    
End Sub
"""

print("Adding macros and pivot...")
wb = xw.Book(output_location)
wb.api.VBProject.VBComponents.Add(1).CodeModule.AddFromString(vba_code)
wb.save(output_location)

macro = wb.macro('Macro2')
macro()

wb.save()
wb.close()

print("Completed..")
t2 = time.time()
print(t2-t1)

os.startfile(output_location)
